import pandas as pd
from decimal import Decimal
from openpyxl.styles import PatternFill, NamedStyle, Alignment, Font, Border, Side, Color
from datetime import datetime
from numpy import std
import tkinter as tk
from tkinter import filedialog
from os import path

# Def colors

colorPairs = [(Color(rgb='ed7d31'),Color(rgb='ffffff')), # orange, white
              (Color(rgb='ffeb9c'),Color(rgb='9c5700')), # light yellow, dark oragne
              (Color(rgb='c6efce'),Color(rgb='006100')), # dark green, light green
              (Color(rgb='a5a5a5'),Color(rgb='ffffff')), # gray, white
              (Color(rgb='FFCC99'),Color(rgb='3f3f76')), # light oragne, purple
              (Color(rgb='ffc7ce'),Color(rgb='9c0006')), # light gray, dark red
              (Color(rgb='990000'),Color(rgb='FF5B00')), # red, orange ✗
              (Color(rgb='FFEE63'),Color(rgb='0e1111')), # yellow, black ◯
              (Color(rgb='D4D925'),Color(rgb='0e1111')) # green, black ✓
              ] 

# Helper functions

def selectExportFile():
    global file1_path
    global expirimentName
    file1_path = filedialog.askopenfilename()
    expirimentName = path.basename(file1_path)
    expirimentName = expirimentName.replace(" Export.xlsx",'')
    
    lbl_file1_path.config(text=f"Export file: {file1_path}")
    root.update_idletasks()  # Update the GUI

def selectLayoutFile():
    global file2_path
    file2_path = filedialog.askopenfilename()
    lbl_file2_path.config(text=f"Layout file: {file2_path}")
    root.update_idletasks()  # Update the GUI
    
def update_tech_name(*args):
    global tech_name
    tech_name = tech_name_var.get()
    
def exportFile():
    if 'file1_path' in globals() and 'file2_path' in globals():
        global export_path, expirimentName
        export_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"{expirimentName} Analysis.xlsx"
        )
        exportFunction()
    else:
        print("Please select both files before exporting.")
        
def center_special_characters(ws):
    # Mapping of characters to color pairs
    char_to_color = {
        '✓': colorPairs[8],
        '◯': colorPairs[7],
        '✗': colorPairs[6]
    }
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in char_to_color:
                fill_color, font_color = char_to_color[cell.value]
                
                # Apply alignment
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Apply fill color
                cell.fill = PatternFill(start_color=fill_color.rgb, end_color=fill_color.rgb, fill_type='solid')
                
                # Apply font color
                cell.font = Font(color=font_color.rgb)

def colorPatient(row, worksheet, color):
    color += 1
    
    etkacCell = worksheet.cell(row=row+4, column=4)
    etkacCell.font = Font(bold=True, size=15)
    etkacCell.alignment = Alignment(horizontal='left')
    
    etkacValueCell = worksheet.cell(row=row+5, column=4)
    etkacValueCell.fill = PatternFill(start_color=colorPairs[color][0], end_color=colorPairs[color][0], fill_type='solid')
    etkacValueCell.font = Font(color=colorPairs[color][1])
    
    for i in range(1,6):
        cell = worksheet.cell(row=row, column=i)
        cell.fill = PatternFill(start_color=colorPairs[color][0], end_color=colorPairs[color][0], fill_type='solid')
        cell.font = Font(color=colorPairs[color][1])
        
    for i in range(4,10):
        cell1 = worksheet.cell(row=row+i, column=1)
        cell2 = worksheet.cell(row=row+i, column=2)
        
        cell1.fill = PatternFill(start_color=colorPairs[0][0], end_color=colorPairs[0][0], fill_type='solid')
        # cell2.fill = PatternFill(start_color=colorPairs[color][0], end_color=colorPairs[color][0], fill_type='solid')
        cell1.font = Font(color=colorPairs[0][1])
        # cell2.font = Font(color=colorPairs[color][1])

def exportFunction():
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException

    # Import the data
    global file1_path
    global file2_path
    global export_path
    global tech_name
    currentTime = datetime.now()

    basicData = pd.read_excel(file1_path)
    basicLayout = pd.read_excel(file2_path)

    # Make export datatables
    exportFrame = pd.DataFrame()
    basalExportFrame = pd.DataFrame()
    summaryFrame = pd.DataFrame()

    # Find experiment name and date
    experimentName = basicLayout.columns.values[4].replace("File Name: ",'')
    experimentDate = basicData.iloc[5,1].strftime('%m/%d/%Y')

    # Locate the corrected data
    firstCol = basicData[basicData.columns[0]].values.tolist()
    for i in range(len(firstCol)):
        if(firstCol[i]=="Corrected [340]"):
            startRow = i+3
            break

    # Extract the corrected data
    times = basicData.iloc[startRow:startRow+61, 1].values.tolist()
    values = basicData.iloc[startRow:startRow+61, 2:98].values.tolist()
    values = pd.DataFrame(values).T.values.tolist()

    # Extract well plate layout
    layout = basicLayout.iloc[1:9, 1:14].values.tolist()
    referenceArray = []

    for row in range(8):
        for col in range(12):
            value = layout[row][col]
            if isinstance(value, float) and str(value) != 'nan':
                value = int(value)
            if str(value) not in referenceArray:
                referenceArray.append(str(value))

    try:
        referenceArray.remove('nan')
    except:
        pass

    trisPatients = {}
    tppPatients = {}
    for patient in referenceArray:
        if (patient != 'nan'):
            trisPatients[patient] = []
            tppPatients[patient] = []

    for row in range(4):
        for col in range(12):
            value = layout[row][col]
            if isinstance(value, float) and str(value) != 'nan':
                value = int(value)
            if str(value)!='nan':
                trisPatients[str(value)].append((row,col))

    for row in range(4, 8):
        for col in range(12):
            value = layout[row][col]
            if isinstance(value, float) and str(value) != 'nan':
                value = int(value)
            if str(value)!='nan':
                tppPatients[str(value)].append((row,col))

    # Insert data into table
    data = [[[] for i in range(12)] for j in range(8)]

    for row in range(8):
        for col in range(12):
            data[row][col] = values[row*12+col]

    # Calculate average rate of change
    averageRatesOfChange = [[0 for i in range(12)] for j in range(8)]

    for row in range(8):
        for col in range(12):
            totalChange = Decimal(0)
            numIntervals = Decimal(20)

            for i in range(40, 60):
                totalChange += Decimal(str(data[row][col][i + 1])) - Decimal(str(data[row][col][i]))

            averageRatesOfChange[row][col] = totalChange/numIntervals

    # Connect data to patients
    trisData = {}
    tppData = {}

    for patient in referenceArray:
        if (patient != 'nan'):
            trisData[patient] = []
            tppData[patient] = []

    for key in trisPatients.keys():
        for i in trisPatients[key]:
            row, col = i
            trisData[key].append(averageRatesOfChange[row][col])

    for key in tppPatients.keys():
        for i in tppPatients[key]:
            row, col = i
            tppData[key].append(averageRatesOfChange[row][col])

    # Convert Data to grids
    trisBlankAverage = sum(trisData['BLK'])/len(trisData['BLK'])
    tppBlankAverage = sum(tppData['BLK'])/len(tppData['BLK'])
    overallBlankAverage = (sum(trisData['BLK'])+sum(tppData['BLK']))/(len(tppData['BLK'])+len(trisData['BLK']))

    if(overallBlankAverage<=-.0010):
        blankText = 'Invalid Blanks'
    else:
        blankText = ''

    blankAverageDisplayGrid = [['',''],
                            ['Tris Blank Average',str(trisBlankAverage)],
                            ['Tpp Blank Average',str(tppBlankAverage)],
                            ['Overall Blank Average', str(overallBlankAverage), blankText]]

    exportFrame = exportFrame._append(blankAverageDisplayGrid, ignore_index=True)

    summaryHeaderGrid = [['',''],
                         ['Software Version','2.0'],
                         ['Last Updated','8/7/2024'],
                         ['Date of ETKAC Expiriment', experimentDate],
                         ['Data of Analysis', currentTime.strftime('%m/%d/%Y')],
                         ['Tech Name', tech_name],
                         ['',''],
                         ['Export File Path',file1_path],
                         ['Layout File Path',file2_path],
                         ['',''],
                         ['Ideal','✓'],
                         ['Unideal','◯'],
                         ['Deficient','✗'],
                         ['',''],
                         ['Sample','Etkac','Tris CV','Tpp CV']]
    
    summaryFrame = summaryFrame._append(summaryHeaderGrid, ignore_index=True)

    referenceArray.remove('BLK')
    referenceArray.sort()

    sufficientCount = 0
    insufficientCount = 0
    deficentCount = 0
    patientCount = 0

    for patient in referenceArray:
        patientCount += 1
        patientNormalizedTrisAvg = sum(trisData[patient])/len(trisData[patient])-overallBlankAverage
        patientTrisStdev = std(trisData[patient])
        patientTrisCv = abs((patientTrisStdev/patientNormalizedTrisAvg)*100)

        patientNormalizedTppAvg = sum(tppData[patient])/len(tppData[patient])-overallBlankAverage
        patientTppStdev = std(tppData[patient])
        patientTppCv = abs((patientTppStdev/patientNormalizedTppAvg)*100)
        
        patientSummaryGrid = [[patient]]

        if(patientNormalizedTppAvg/patientNormalizedTrisAvg >= 1.25):
            etkacText = 'Deficent'
            patientSummaryGrid[0].append('✗')
            deficentCount += 1
        elif(patientNormalizedTppAvg/patientNormalizedTrisAvg >= 1.15):
            etkacText = 'Insufficient'
            patientSummaryGrid[0].append('◯')
            insufficientCount += 1
        else:
            etkacText = ''
            patientSummaryGrid[0].append('✓')
            sufficientCount += 1

        if(patientTrisCv > 6):
            trisCvText = 'High'
            patientSummaryGrid[0].append('✗')
        elif(patientTrisCv > 5):
            trisCvText = 'Unideal'
            patientSummaryGrid[0].append('◯')
        else:
            trisCvText = ''
            patientSummaryGrid[0].append('✓')

        if(patientTppCv > 6):
            tppCvText = 'High'
            patientSummaryGrid[0].append('✗')
        elif(patientTppCv > 5):
            tppCvText = 'Unideal'
            patientSummaryGrid[0].append('◯')
        else:
            tppCvText = ''
            patientSummaryGrid[0].append('✓')

        basalGrid = [[patient, round(abs(patientNormalizedTrisAvg),4), '', f"=(B{patientCount+1}*1404.4501)/C{patientCount+1}"]]
        basalExportFrame = basalExportFrame._append(basalGrid, ignore_index=True)
        summaryFrame = summaryFrame._append(patientSummaryGrid, ignore_index=True)

        patientGrid = [['',''],
                    ['NORMALIZED TO OVERALL BLANK', patient, patient, patient, patient],
                    ['TRIS'],
                    ['TPP'],
                    ['',''],
                    ['Average TRIS', round(patientNormalizedTrisAvg,11),'','ETKAC'],
                    ['STD TRIS', round(patientTrisStdev,11), '', round(patientNormalizedTppAvg/patientNormalizedTrisAvg,11),etkacText],
                    ['CV% TRIS', round(patientTrisCv,11), trisCvText],
                    ['Average TPP', round(patientNormalizedTppAvg,11)],
                    ['STD TPP', round(patientTppStdev,11)],
                    ['CV% TPP', round(patientTppCv,11), tppCvText],
                    ['']]
                
        for item in trisData[patient]:
            patientGrid[2].append(item-overallBlankAverage)
        for item in tppData[patient]:
            patientGrid[3].append(item-overallBlankAverage)

        exportFrame = exportFrame._append(patientGrid, ignore_index=True)   

    lbl_analysis_report.config(text = f"Sufficient: {sufficientCount} | Insufficient: {insufficientCount} | Deficent: {deficentCount}") 

    root.update_idletasks()  # Update the GUI

    # Exporting to excel file
    summaryFrame = summaryFrame.rename(columns={exportFrame.columns[0]: experimentName, exportFrame.columns[1]: experimentDate, exportFrame.columns[2]: '', exportFrame.columns[3]: '', exportFrame.columns[4]: ''})
    exportFrame = exportFrame.rename(columns={exportFrame.columns[0]: experimentName, exportFrame.columns[1]: experimentDate, exportFrame.columns[2]: '', exportFrame.columns[3]: '', exportFrame.columns[4]: ''})
    basalExportFrame = basalExportFrame.rename(columns={basalExportFrame.columns[0]: "Sample", basalExportFrame.columns[1]: "Avg Tris", basalExportFrame.columns[2]: "Hgb g/DL", basalExportFrame.columns[3]: "Basal Activity"})
    
    try:
        with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
            summaryFrame.to_excel(writer, sheet_name="Summary", index=False)
            exportFrame.to_excel(writer, sheet_name="Etkac", index=False)
            basalExportFrame.to_excel(writer, sheet_name="Basal Activity", index=False)

            worksheet = writer.sheets["Etkac"]
            basalWorksheet = writer.sheets["Basal Activity"]
            summaryWorksheet = writer.sheets["Summary"]

            # Stylize excel file
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = (max_length + 2) * 1.0
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
            for column in basalWorksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = (max_length + 2) * 1.0
                basalWorksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
            summaryWorksheet.column_dimensions['A'].width = 24
            summaryWorksheet.column_dimensions['B'].width = 12
            summaryWorksheet.column_dimensions['C'].width = 12
            summaryWorksheet.column_dimensions['D'].width = 12
            
            center_special_characters(summaryWorksheet)
            
            worksheet['a1'].border = None
            worksheet['b1'].border = None
            worksheet['c1'].border = None
            worksheet['d1'].border = None
            worksheet['e1'].border = None
            
            basalWorksheet['a1'].border = None
            basalWorksheet['b1'].border = None
            basalWorksheet['c1'].border = None
            basalWorksheet['d1'].border = None
            
            summaryWorksheet['a1'].border = None
            summaryWorksheet['b1'].border = None
            summaryWorksheet['c1'].border = None
            summaryWorksheet['d1'].border = None
            

            # Blank flagging
            if(float(exportFrame.iloc[3,1])<=-.0010):
                worksheet.cell(row=5, column=2).fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')

            # Stylize all Patients
            for i in range(len(referenceArray)):
                # Color Patient Header
                colorPatient((i*12)+7, worksheet, i%5)

    except InvalidFileException as e:
        lbl_analysis_report.config(text=f"Error: Cannot write to file as it is already open. Please close the file and try again.")
        print(f"Error: Cannot write to file as it is already open. Please close the file and try again. Exception: {e}")

    except PermissionError as e:
        lbl_analysis_report.config(text=f"Error: Permission denied. The file might be open in another program. Please close the file and try again.")
        print(f"Error: Permission denied. The file might be open in another program. Please close the file and try again. Exception: {e}")

    except Exception as e:
        lbl_analysis_report.config(text=f"An unexpected error occurred: {e}")
        print(f"An unexpected error occurred: {e}")
   
# Setup Tkinter Gui

root = tk.Tk()
root.title("ETKAC Analysis program")
root.geometry("350x1")  # Set initial width to 350 and a minimal height
root.minsize(500, 500)    # Set minimum width to 350 and minimum height to 1

def adjust_window_height(event=None):
    root.update_idletasks()
    root.geometry(f"350x{root.winfo_reqheight()}")

# Create entry for tech name
tech_name_var = tk.StringVar()
tech_name_var.trace("w", update_tech_name)
lbl_tech_name_entry = tk.Label(root, text="Tech Name:")
entry_tech_name = tk.Entry(root, textvariable=tech_name_var)

# Create buttons with padding
btn_file1 = tk.Button(root, text="Select Export File", command=selectExportFile)
btn_file2 = tk.Button(root, text="Select Layout File", command=selectLayoutFile)
btn_export = tk.Button(root, text="Create Analysis Report", command=exportFile)

# Create labels for file paths and the analysis report
lbl_file1_path = tk.Label(root, text="Export file: Not selected", wraplength=300, justify="center", bg="light gray", padx=5, pady=5)
lbl_file2_path = tk.Label(root, text="Layout file: Not selected", wraplength=300, justify="center", bg="light gray", padx=5, pady=5)
lbl_analysis_report = tk.Label(root, text="", wraplength=300, justify="center", bg="light gray", padx=5, pady=5)

# Place widgets in the window
lbl_tech_name_entry.pack(pady=(10,0))
entry_tech_name.pack(pady=(0,10))

# Pack buttons with horizontal padding
btn_file1.pack(pady=5, padx=50)
lbl_file1_path.pack(pady=5, fill="x", padx=10)
btn_file2.pack(pady=5, padx=50)
lbl_file2_path.pack(pady=5, fill="x", padx=10)
btn_export.pack(pady=5, padx=50)
lbl_analysis_report.pack(pady=5, fill="x", padx=10)

# Bind the adjust_window_height function to configure events of all widgets
for widget in root.winfo_children():
    widget.bind('<Configure>', adjust_window_height)

# Initial adjustment of window height
adjust_window_height()

# Start the GUI event loop
root.mainloop()